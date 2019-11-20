import os
import sys
from openpyxl import load_workbook, Workbook
import Main

remove_keypoint_num = 10 #눈,귀,발 제외(15~24)
Lunge_CompData_path = os.path.dirname( os.path.abspath( __file__ ) )+'/Data/Lunge_CompData.xlsx'
Squat_CompData_path = os.path.dirname( os.path.abspath( __file__ ) )+'./Data/Squat_CompData.xlsx'
Flank_CompData_path = os.path.dirname( os.path.abspath( __file__ ) )+'./Data/Flank_CompData.xlsx'

def merge_JsonFile(dir_path):
    if (os.path.exists(dir_path)!=True):
        print("Not found "+dir_path+" directory!!!")
        sys.exit(0)
    file_list = os.listdir(dir_path)
    file_list.sort()

    write_wb = Workbook()
    write_ws = write_wb.active

    count = 0
    for i in file_list:
        fname, ext = os.path.splitext(i)
        if(ext=='.json'):
            count+=1
            f= open(dir_path+'/'+i)
            keypoint = f.readline()
            keypoint = keypoint[keypoint.find('pose_keypoints_2d":[') + 20:keypoint.find('],"face_keypoints_2d"')]
            keypoint = keypoint.split(',')
            for j in range(len(keypoint)):
                keypoint[j]=float(keypoint[j])
                write_ws.cell(row=j+1, column = count).value = keypoint[j]

    dir_name = dir_path.split(os.path.sep)
    dir_name=dir_name[-1]
    save_name = dir_path+'/'+dir_name +'.xlsx'
    write_wb.save(save_name)


def get_data(file_path, m=0):#m==0:input data, #m==1:comparisonData
    if (os.path.exists(file_path) != True):
        print("Not found '" + file_path + "!!!")
        return -1

    load_wb = load_workbook(file_path, data_only=True)
    load_ws = load_wb.active

    if(m==0):
        all_values = [[[] for col in range(int(load_ws.max_row / 3))] for row in range(3)]
        for i in range(load_ws.max_row):#0~74
            row = load_ws[i+1]
            for cell in row:
                # if(float(cell.value)<0):
                #     continue
                if(i%3==0):#x
                    all_values[0][int(i/3)].append(float(cell.value))
                elif(i%3==1):#y
                    all_values[1][int(i/3)].append(float(cell.value))
                else:#c
                    all_values[2][int(i/3)].append(float(cell.value))


    else:#m==1(comp_data)
        all_values = [[[] for col in range(int(load_ws.max_row / 2))] for row in range(2)]
        for i in range(load_ws.max_row):  # 0~74
            row = load_ws[i + 1]
            for cell in row:
                if (float(cell.value) < 0):
                    continue
                if (i % 2 == 0):  # x
                    all_values[0][int(i / 2)].append(float(cell.value))
                elif (i % 2 == 1):  # y
                    all_values[1][int(i / 2)].append(float(cell.value))

    load_wb.close()
    return all_values

def refine_data(_3d_list, save_name):

    write_wb = Workbook()
    write_ws = write_wb.active

    all_x_keypoint = [[] for col in range(Main.keypoint_num)]
    all_y_keypoint = [[] for col in range(Main.keypoint_num)]

    for i in range(Main.keypoint_num):#_3d_list[0]:x, [1]:y, [2]:c
        for j in range(len(_3d_list[0][i])):
            if (_3d_list[2][i][j] >= 0.2):#c가 0.2이상
                all_x_keypoint[i].append(_3d_list[0][i][j])
                all_y_keypoint[i].append(_3d_list[1][i][j])
                write_ws.cell(row=(i*2)+1, column = j+1).value = _3d_list[0][i][j]
                write_ws.cell(row=(i+1)*2, column = j+1).value = _3d_list[1][i][j]
            else:
                write_ws.cell(row=(i*2)+1, column=j+1).value = -1
                write_ws.cell(row=(i+1)*2, column=j+1).value = -1

    write_wb.save(save_name)
    return all_x_keypoint, all_y_keypoint

def save_reduceList_toExcel(x_list, y_list):
    write_wb = Workbook()
    write_ws = write_wb.active

    for i in range(Main.keypoint_num):
        for j in range(len(x_list[i])):
            write_ws.cell(row=(i * 2) + 1, column=j + 1).value = x_list[i][j]
        for j in range(len(y_list[i])):
            write_ws.cell(row=(i + 1) * 2, column=j + 1).value = y_list[i][j]

    print(os.path.dirname(os.getcwd()))
    write_wb.save('reduece_' + os.path.dirname(os.getcwd()).split(os.path.sep)[-1]+'.xlsx')


# if __name__ == '__main__':
#     dir_path = input('파일 경로 입력 >> ')
#     merge_JsonFile(dir_path)
#     s = os.path.split(dir_path)
#     excel_saveName = s[len(s) - 1] + '.xlsx'
#     all_values = get_data(dir_path + '/' + excel_saveName, 0)
#     refine_data(all_values, dir_path+'/'+excel_saveName)